# %%
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from tqdm import tqdm
from datetime import datetime

from crisp.library import start
from crisp.library import secrets
from crisp.library import format_prompts
from crisp.library import classify
from crisp.library import metric_standard_errors

# ------------------ CONSTANTS ------------------
OPENAI_API_KEY = secrets.OPENAI_API_KEY
MODEL = start.MODEL
GENERATION_TEMPERATURE = 0.00001
NUM_RESPONSES = 5
SEED = start.SEED
DEV_OUTPUT_PATH = start.DATA_DIR + "temp/ncb_persona_responses_dev.xlsx"
DESTINATION_PATH = start.MAIN_DIR + "results/ncb_persona_results_dev.xlsx"
PERSONA_RESULTS_PATH = start.DATA_DIR + "temp/ncb_persona_training_results.csv"

PROMPT_PATH_BEST = start.DATA_DIR + "temp/ncb_ape_best.csv"
PROMPT_PATH_WORST = start.DATA_DIR + "temp/ncb_ape_worst.csv"

client = OpenAI(api_key=OPENAI_API_KEY)

# ------------------ LOAD BEST PERSONAS ------------------
persona_df = pd.read_csv(PERSONA_RESULTS_PATH)
best_persona = (
    persona_df[persona_df.prompt_type == "best"]
    .sort_values("f1_score", ascending=False)
    .iloc[0]["persona"]
)
worst_persona = (
    persona_df[persona_df.prompt_type == "worst"]
    .sort_values("f1_score", ascending=False)
    .iloc[0]["persona"]
)
print(f"Using best persona for BEST prompt: {best_persona}")
print(f"Using best persona for WORST prompt: {worst_persona}")

# ------------------ LOAD PROMPTS ------------------
prompt_df1 = pd.read_csv(PROMPT_PATH_BEST)
prompt_df1["prompt_id"] = (
    "best" + prompt_df1.generation.astype(str) + "_" + prompt_df1.variant_id.astype(str)
)
prompt_df1 = prompt_df1.set_index("prompt_id")
max_f1_index1 = prompt_df1["f1_score"].idxmax()
prompt_df1 = prompt_df1.loc[[max_f1_index1]]
prompt_df1["persona"] = best_persona

prompt_df2 = pd.read_csv(PROMPT_PATH_WORST)
prompt_df2["prompt_id"] = (
    "worst"
    + prompt_df2.generation.astype(str)
    + "_"
    + prompt_df2.variant_id.astype(str)
)
prompt_df2 = prompt_df2.set_index("prompt_id")
max_f1_index2 = prompt_df2["f1_score"].idxmax()
prompt_df2 = prompt_df2.loc[[max_f1_index2]]
prompt_df2["persona"] = worst_persona

prompt_df = pd.concat([prompt_df1, prompt_df2])
prompt_df["text"] = prompt_df["persona"] + "\n" + prompt_df["prompt"]
prompt_df["combo_id"] = prompt_df.index

# ------------------ FORMAT PROMPTS ------------------
formatted_prompts = []
prompt_mapping = {}
for combo_id, prompt_text in zip(prompt_df.combo_id, prompt_df.text):
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])
    prompt_mapping[prompt_text] = combo_id

# ------------------ LOAD TEXT DATA ------------------
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "dev"]

# ------------------ COLLECT RESPONSES ------------------
response_rows = []
for loop_num in [1, 2]:
    for prompt in formatted_prompts:
        prompt_text = prompt[0]["content"]
        combo_id = prompt_mapping[prompt_text]

        for text, participant_id, study, question, human_code in tqdm(
            zip(df.text, df.participant_id, df.study, df.question, df.human_code),
            total=len(df),
        ):
            timestamp = datetime.now().isoformat()
            messages = prompt + [{"role": "user", "content": text}]
            response = client.chat.completions.create(
                model=MODEL,
                messages=messages,
                temperature=GENERATION_TEMPERATURE,
                n=NUM_RESPONSES,
                seed=SEED,
            )

            for i, choice in enumerate(response.choices):
                cleaned_response = choice.message.content
                classification = classify.create_binary_classification_from_response(
                    cleaned_response
                )

                response_rows.append(
                    {
                        "participant_id": participant_id,
                        "study": study,
                        "question": question,
                        "text": text,
                        "human_code": human_code,
                        "response": cleaned_response,
                        "classification": classification,
                        "prompt": prompt_text,
                        "model": MODEL,
                        "fingerprint": response.system_fingerprint,
                        "combo_id": combo_id,
                        "response_number": i + 1,
                        "loop_num": loop_num,
                        "timestamp": timestamp,
                    }
                )

long_df = pd.DataFrame(response_rows)
long_df.to_excel(DEV_OUTPUT_PATH, index=False)

# ------------------ EXPORT METRICS ------------------
wb = load_workbook(DESTINATION_PATH)
wb.create_sheet("combo_results")
ws = wb["combo_results"]

headers = [
    "Combo",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Accuracy SE",
    "Precision SE",
    "Recall SE",
    "F1 SE",
    "Prompt",
]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[
        (long_df.combo_id == combo)
        & (long_df.response_number == 1)
        & (long_df.loop_num == 1)
    ]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]

    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    _, accuracy_se = metric_standard_errors.bootstrap_accuracy(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, precision_se = metric_standard_errors.bootstrap_precision(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, recall_se = metric_standard_errors.bootstrap_recall(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, f1_se = metric_standard_errors.bootstrap_f1(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )

    results = [
        combo,
        accuracy,
        precision,
        recall,
        f1,
        accuracy_se,
        precision_se,
        recall_se,
        f1_se,
        combo_df.prompt.iloc[0],
    ]
    for col, val in enumerate(results, 1):
        ws.cell(
            row=row, column=col, value=round(val, 2) if isinstance(val, float) else val
        )

    row += 1

wb.save(DESTINATION_PATH)
# %%

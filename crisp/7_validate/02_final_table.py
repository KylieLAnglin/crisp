# %%
import os
import pandas as pd
from crisp.library import start
from openpyxl import load_workbook

RESULTS_FILE = start.MAIN_DIR + "results/final_results.xlsx"

# %%

gratitude_df = pd.read_excel(
    start.MAIN_DIR + "results/gratitude_best_results_test.xlsx", sheet_name="results"
)
ncb_df = pd.read_excel(
    start.MAIN_DIR + "results/ncb_best_results_test.xlsx", sheet_name="results"
)
mm_df = pd.read_excel(
    start.MAIN_DIR + "results/mm_best_results_test.xlsx", sheet_name="results"
)
# %%

wb = load_workbook(RESULTS_FILE)
ws = wb.active

# one row/outcome at a time
row = 2
for concept, df in zip(["gratitude", "ncb", "mm"], [gratitude_df, ncb_df, mm_df]):
    col = 2
    ws.cell(row=row, column=1, value=concept)
    for metric in ["Accuracy", "Precision", "Recall", "F1"]:
        metric_value = round(df[metric].values[0], 2)
        ws.cell(row=row, column=col, value=metric_value)
        col += 1
    col = 2
    row += 1

    for metric in ["Accuracy", "Precision", "Recall", "F1"]:
        se_value = round(df[metric + " SE"].values[0], 2)
        se_formatted = f"({se_value})"
        ws.cell(row=row, column=col, value=se_formatted)
        col += 1
    row += 1

wb.save(RESULTS_FILE)

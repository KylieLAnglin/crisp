# %%
# 01_baseline_train.py
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
from datetime import datetime
import os

from crisp.library import start, secrets
from crisp.library import format_prompts, classify

CONCEPT = start.CONCEPT
PLATFORM = start.PLATFORM
MODEL = start.MODEL
print(f"Running {CONCEPT} on {PLATFORM} with {MODEL} in train set")
# ------------------ CONSTANTS ------------------

IMPORT_PROMPT_PATH = start.DATA_DIR + f"prompts/{CONCEPT}_baseline_variants.xlsx"
DATA_PATH = start.DATA_DIR + f"clean/{CONCEPT}.xlsx"

EXPORT_RESPONSE_PATH = (
    start.DATA_DIR
    + f"responses_train/{PLATFORM}_{CONCEPT}_baseline_zero_responses_train.xlsx"
)
EXPORT_RESULTS_PATH = (
    start.MAIN_DIR + f"results/{PLATFORM}_{CONCEPT}_baseline_zero_results_train.xlsx"
)


# ------------------ LOAD DATA ------------------
df = pd.read_excel(DATA_PATH)
df = df[df.split_group == "train"]
if start.SAMPLE:
    df = df.sample(5, random_state=start.SEED)

prompt_df = pd.read_excel(IMPORT_PROMPT_PATH, sheet_name="baseline")
prompt_df["baseline_prompt_id"] = prompt_df.index
full_prompt_texts = prompt_df["prompt"].tolist()

# TODO: does this need to be edited for llama? If so,
# add an argument for platform
formatted_prompts = []
for prompt_text in full_prompt_texts:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])

# %%
# ------------------ COLLECT RESPONSES ------------------
response_rows = []
for baseline_prompt_id, prompt in zip(prompt_df.baseline_prompt_id, formatted_prompts):
    prompt_text = prompt[0]["content"]  # TODO: is this the same format for llama?

    for text, participant_id, study, question, human_code in tqdm(
        zip(df.text, df.participant_id, df.study, df.question, df.human_code),
        total=len(df),
        desc=f"Prompt ID: {baseline_prompt_id}",
    ):
        timestamp = datetime.now().isoformat()
        cleaned_response, system_fingerprint = classify.format_message_and_get_response(
            model_provider=start.PLATFORM,
            prompt=prompt,
            text_to_classify=text,
            temperature=0.0001,
        )
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )

        response_rows.append(
            {
                "participant_id": participant_id,
                "study": study,
                "question": question,
                "text": text,
                "human_code": human_code,
                "response": cleaned_response,
                "classification": classification,
                "prompt": prompt_text,
                "model": start.MODEL,
                "fingerprint": system_fingerprint,
                "baseline_prompt_id": baseline_prompt_id,
                "timestamp": timestamp,
            }
        )

long_df = pd.DataFrame(response_rows)
long_df.to_excel(EXPORT_RESPONSE_PATH, index=False)

# ------------------ CREATE DESTINATION FILE ------------------
if not os.path.exists(EXPORT_RESULTS_PATH):
    wb = Workbook()
    wb.save(EXPORT_RESULTS_PATH)

# ------------------ EVALUATE PROMPTS ------------------
wb = load_workbook(EXPORT_RESULTS_PATH)
if "results" in wb.sheetnames:
    del wb["results"]
ws = wb.create_sheet("results")

headers = [
    "Baseline Prompt ID",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Prompt",
]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

row = 2
for bp_id in long_df.baseline_prompt_id.unique():
    combo_df = long_df[long_df.baseline_prompt_id == bp_id]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )

    results = [
        bp_id,
        accuracy,
        precision,
        recall,
        f1,
        combo_df.prompt.iloc[0],
    ]
    for col, val in enumerate(results, 1):
        ws.cell(
            row=row, column=col, value=round(val, 3) if isinstance(val, float) else val
        )
    row += 1

wb.save(EXPORT_RESULTS_PATH)
# %%

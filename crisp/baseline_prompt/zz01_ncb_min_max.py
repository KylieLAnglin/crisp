# %%
import shutil
import os

import numpy as np
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from tqdm import tqdm


from crisp.library import start
from crisp.library import secrets
from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify


PROMPT_FILE = "ncb_min_max_prompt.xlsx"
source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE
destination_path = start.MAIN_DIR + "results/" + PROMPT_FILE

df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "train"]

# %% Import prompts and create system messages
prompt_components = pd.read_excel(source_path, sheet_name="components")
prompt_components = prompt_components.set_index("component")

prompt_combos = pd.read_excel(source_path, sheet_name="combos")
prompt_combos = prompt_combos.set_index("combo_id")

system_message_dicts = format_prompts.combine_prompt_text_components(
    prompt_components, prompt_combos
)
system_message_dfs = pd.DataFrame(system_message_dicts)
# %% Format prompts
formatted_prompts = []
for i, row in system_message_dfs.iterrows():
    formatted_prompts.append(format_prompts.format_system_message(row["text"]))

# %% Get responses
response_dfs = []
combo_id = 1
for prompt in formatted_prompts:
    responses = []
    classifications = []
    for text in tqdm(df.text):
        cleaned_response = classify.format_message_and_get_response(
            [prompt], text, model="gpt-4o", temperature=0.00
        )
        responses.append(cleaned_response)
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )
        classifications.append(classification)
    response_df = pd.DataFrame(
        {
            "participant_id": df.participant_id,
            "study": df.study,
            "question": df.question,
            "text": df.text,
            "human_code": df.human_code,
            "response": responses,
            "classification": classifications,
            "combo_id": combo_id,
        }
    )
    response_dfs.append(response_df)
    combo_id += 1

long_df = pd.concat(response_dfs)
long_df = long_df.dropna(subset=["classification"])

# %%
export_results.copy_file_and_paste(source_path, destination_path)
wb = load_workbook(destination_path)
ws = wb["combos"]
previous_column_count = ws.max_column
export_results.prep_metrics_workbook(destination_path, wb, ws)

for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    export_results.export_metrics(
        destination_path=destination_path,
        wb=wb,
        ws=ws,
        row=combo + 1,
        col=previous_column_count + 1,
        accuracy=accuracy,
        precision=precision,
        recall=recall,
        f1=f1,
    )

# %%

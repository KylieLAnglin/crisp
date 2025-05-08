###
# Tests five definitions and five tasks.
# Print prompts and performance to destination_path below
###
# %%
from crisp.library import start

PROMPT_FILE = "ncb_variants"
# import prompts from here
source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE + ".xlsx"
# export results here, prompts are copied to this file
destination_path = start.MAIN_DIR + "results/" + PROMPT_FILE + "_results.xlsx"
export_results.copy_file_and_paste(source_path, destination_path)

# %%

import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from tqdm import tqdm
from datetime import datetime


from crisp.library import start
from crisp.library import secrets
from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify

OPENAI_API_KEY = secrets.OPENAI_API_KEY

client = OpenAI(api_key=OPENAI_API_KEY)

# text data
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "train"]

# %% Import prompts and create combos
part1_variants = pd.read_excel(source_path, sheet_name="part1", index_col="part_num")
part2_variants = pd.read_excel(source_path, sheet_name="part2", index_col="part_num")

prompt_combos = [
    {
        "text": f"{part1} {part2} ",  # Combine prompt text
        "part1": part1_id,  # Store Part 1 ID
        "part2": part2_id,  # Store Part 2 ID
    }
    for part1_id, part1 in part1_variants["prompt_part"].items()  # Iterate over part1
    for part2_id, part2 in part2_variants[
        "prompt_part"
    ].items()  # Nested: iterate over part2
]
prompt_df = pd.DataFrame(prompt_combos)
prompt_df["combo_id"] = prompt_df.index

full_prompt_texts = prompt_df.text.tolist()
# %%

formatted_prompts = []
for prompt_text in full_prompt_texts:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])


# %%
# Test response
prompt = formatted_prompts[1]
text = df.text.iloc[0]
messages = prompt + [{"role": "user", "content": text}]
response = client.chat.completions.create(
    model=start.MODEL, messages=messages, temperature=0.00001, n=1, seed=start.SEED
)
# print fingerprint
print(response.system_fingerprint)
for choice in response.choices:
    print(choice.message.content)
#

# %% Get responses (one row per response, with timestamp)
response_rows = []
for loop_num in [1, 2]:
    for prompt in formatted_prompts:
        prompt_text = prompt[0]["content"]
        combo_id = prompt_df[prompt_df.text == prompt_text].combo_id.iloc[0]

        for text, participant_id, study, question, human_code in tqdm(
            zip(df.text, df.participant_id, df.study, df.question, df.human_code),
            total=len(df),
        ):
            timestamp = datetime.now().isoformat()
            messages = prompt + [{"role": "user", "content": text}]
            response = client.chat.completions.create(
                model=start.MODEL,
                messages=messages,
                temperature=0.00001,
                n=5,
                seed=start.SEED,
            )

            for i, choice in enumerate(response.choices):
                cleaned_response = choice.message.content
                classification = classify.create_binary_classification_from_response(
                    cleaned_response
                )

                response_rows.append(
                    {
                        "participant_id": participant_id,
                        "study": study,
                        "question": question,
                        "text": text,
                        "human_code": human_code,
                        "response": cleaned_response,
                        "classification": classification,
                        "prompt": prompt_text,
                        "model": start.MODEL,
                        "fingerprint": response.system_fingerprint,
                        "combo_id": combo_id,
                        "response_number": i + 1,
                        "timestamp": timestamp,
                    }
                )

long_df = pd.DataFrame(response_rows)
long_df.to_excel(start.DATA_DIR + "temp/ncb_variants_responses.xlsx", index=False)
# %%
# Load empty results file (just has prompts sheets)
wb = load_workbook(destination_path)
wb.create_sheet("combo_results")
ws = wb["combo_results"]

ws.cell(row=1, column=1, value="Combo")
ws.cell(row=1, column=2, value="Part 1 ID")
ws.cell(row=1, column=3, value="Part 2 ID")
ws.cell(row=1, column=4, value="Accuracy")
ws.cell(row=1, column=5, value="Precision")
ws.cell(row=1, column=6, value="Recall")
ws.cell(row=1, column=7, value="F1")
ws.cell(row=1, column=8, value="Prompt")

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    part1_id = prompt_df.loc[prompt_df.combo_id == combo, "part1"].values[0]
    part2_id = prompt_df.loc[prompt_df.combo_id == combo, "part2"].values[0]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    ws.cell(row=row, column=1, value=combo)
    ws.cell(row=row, column=2, value=part1_id)
    ws.cell(row=row, column=3, value=part2_id)
    ws.cell(row=row, column=4, value=round(accuracy, 3))
    ws.cell(row=row, column=5, value=round(precision, 3))
    ws.cell(row=row, column=6, value=round(recall, 3))
    ws.cell(row=row, column=7, value=round(f1, 3))
    ws.cell(row=row, column=8, value=combo_df.prompt.iloc[0])
    row = row + 1

    wb.save(destination_path)

# %%

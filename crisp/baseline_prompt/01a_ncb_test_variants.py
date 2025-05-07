###
# Tests three definitions and three tasks.
# Print prompts and performance to destination_path below
###
# %%
from crisp.library import start

PROMPT_FILE = "ncb_variants"
# import prompts from here
source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE + ".xlsx"
# export results here, prompts are copied to this file
destination_path = start.MAIN_DIR + "results/" + PROMPT_FILE + "_results.xlsx"

# %%

import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from tqdm import tqdm


from crisp.library import start
from crisp.library import secrets
from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify

OPENAI_API_KEY = secrets.OPENAI_API_KEY

client = OpenAI(api_key=OPENAI_API_KEY)

# text data
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "train"]
# prompts and prep results file
export_results.copy_file_and_paste(
    source_path, destination_path
)  # copy prompts to results folder

# %% Import prompts
part1_variants = pd.read_excel(source_path, sheet_name="part1")
part1_variants = part1_variants.set_index("part_num")
part2_variants = pd.read_excel(source_path, sheet_name="part2")
part2_variants = part2_variants.set_index("part_num")

# %%
# Create combos
full_prompt_texts = []
part1_ids = []
part2_ids = []
for part1, part1_id in zip(part1_variants["prompt_part"], part1_variants.index):
    for part2, part2_id in zip(part2_variants["prompt_part"], part2_variants.index):
        full_prompt_texts.append(part1 + " " + part2 + " ")
        part1_ids.append(part1_id)
        part2_ids.append(part2_id)


prompt_df = pd.DataFrame(full_prompt_texts, columns=["text"])
prompt_df["part1"] = part1_ids
prompt_df["part2"] = part2_ids
prompt_df["combo_id"] = prompt_df.index

# %%

formatted_prompts = []
for prompt_text in full_prompt_texts:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])


# %%
# Test response
prompt = formatted_prompts[1]
text = df.text.iloc[0]
messages = prompt + [{"role": "user", "content": text}]
response = client.chat.completions.create(
    model=start.MODEL, messages=messages, temperature=0.0001, n=1, seed=start.SEED
)
# print fingerprint
print(response.system_fingerprint)
for choice in response.choices:
    print(choice.message.content)
# %%
# %% Get responses
response_dfs = []
for prompt in formatted_prompts:
    responses1 = []
    responses2 = []
    responses3 = []
    classifications1 = []
    classifications2 = []
    classifications3 = []
    fingerprints = []

    for text in tqdm(df.text):
        messages = prompt + [{"role": "user", "content": text}]

        response = client.chat.completions.create(
            model=start.MODEL,
            messages=messages,
            temperature=0.0001,
            n=3,
            seed=start.SEED,
        )

        cleaned_response1 = response.choices[0].message.content
        cleaned_response2 = response.choices[1].message.content
        cleaned_response3 = response.choices[2].message.content
        classification1 = classify.create_binary_classification_from_response(
            cleaned_response1
        )
        classification2 = classify.create_binary_classification_from_response(
            cleaned_response2
        )
        classification3 = classify.create_binary_classification_from_response(
            cleaned_response3
        )
        responses1.append(cleaned_response1)
        responses2.append(cleaned_response2)
        responses3.append(cleaned_response3)
        classifications1.append(classification1)
        classifications2.append(classification2)
        classifications3.append(classification3)
        fingerprints.append(response.system_fingerprint)
    response_df = pd.DataFrame(
        {
            "participant_id": df.participant_id,
            "study": df.study,
            "question": df.question,
            "text": df.text,
            "human_code": df.human_code,
            "response1": responses1,
            "classification1": classifications1,
            "prompt": prompt[0]["content"],
            "model": start.MODEL,
            "fingerprint": fingerprints,
            "combo_id": prompt_df[prompt_df.text == prompt[0]["content"]].combo_id.iloc[
                0
            ],
            "response2": responses2,
            "classification2": classifications2,
            "response3": responses3,
            "classification3": classifications3,
        }
    )
    response_dfs.append(response_df)

long_df = pd.concat(response_dfs)
# long_df = long_df.dropna(subset=["classification"])
long_df.to_excel(start.DATA_DIR + "temp/ncb_variants_responses.xlsx", index=False)
# %%
# Load empty results file (just has prompts sheets)
wb = load_workbook(destination_path)
wb.create_sheet("combo_results")
ws = wb["combo_results"]

ws.cell(row=1, column=1, value="Combo")
ws.cell(row=1, column=2, value="Part 1 ID")
ws.cell(row=1, column=3, value="Part 2 ID")
ws.cell(row=1, column=4, value="Accuracy")
ws.cell(row=1, column=5, value="Precision")
ws.cell(row=1, column=6, value="Recall")
ws.cell(row=1, column=7, value="F1")
ws.cell(row=1, column=8, value="Prompt")

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    part1_id = prompt_df.loc[prompt_df.combo_id == combo, "part1"].values[0]
    part2_id = prompt_df.loc[prompt_df.combo_id == combo, "part2"].values[0]
    valid_indices = combo_df.dropna(subset=["human_code", "classification1"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification1"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    ws.cell(row=row, column=1, value=combo)
    ws.cell(row=row, column=2, value=part1_id)
    ws.cell(row=row, column=3, value=part2_id)
    ws.cell(row=row, column=4, value=round(accuracy, 2))
    ws.cell(row=row, column=5, value=round(precision, 2))
    ws.cell(row=row, column=6, value=round(recall, 2))
    ws.cell(row=row, column=7, value=round(f1, 2))
    ws.cell(row=row, column=8, value=combo_df.prompt.iloc[0])
    row = row + 1

    wb.save(destination_path)

# %%

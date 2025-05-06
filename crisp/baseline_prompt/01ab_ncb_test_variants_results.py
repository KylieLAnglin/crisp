# %%

import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from tqdm import tqdm
import numpy as np

from crisp.library import start
from crisp.library import secrets
from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify


# FILE = "ncb_variants_responses 2025-05-06 144 pm"
FILE = "ncb_variants_responses 2025-05-06 316 pm"
long_df = pd.read_excel(start.DATA_DIR + "temp/" + FILE + ".xlsx")

# votecount across classification1, classification2, classification3
long_df["classification"] = long_df[
    [
        "classification1",
        "classification2",
        "classification3",
    ]
].mode(axis=1)[0]
# %%
PROMPT_FILE = "ncb_variants"
# import prompts from here
source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE + ".xlsx"
# export results here, prompts are copied to this file
# destination_path = (
#     start.MAIN_DIR + "results/" + PROMPT_FILE + "_results 2025-05-06 144 pm.xlsx"
# )
destination_path = (
    start.MAIN_DIR + "results/" + PROMPT_FILE + "_results 2025-05-06 316 pm.xlsx"
)
export_results.copy_file_and_paste(
    source_path, destination_path
)  # copy prompts to results folder


# %%

# %% Import prompts
part1_variants = pd.read_excel(source_path, sheet_name="part1")
part1_variants = part1_variants.set_index("part_num")
part2_variants = pd.read_excel(source_path, sheet_name="part2")
part2_variants = part2_variants.set_index("part_num")

# %%
# Create combos
full_prompt_texts = []
part1_ids = []
part2_ids = []
for part1, part1_id in zip(part1_variants["prompt_part"], part1_variants.index):
    for part2, part2_id in zip(part2_variants["prompt_part"], part2_variants.index):
        full_prompt_texts.append(part1 + " " + part2 + " ")
        part1_ids.append(part1_id)
        part2_ids.append(part2_id)


prompt_df = pd.DataFrame(full_prompt_texts, columns=["text"])
prompt_df["part1"] = part1_ids
prompt_df["part2"] = part2_ids
prompt_df["combo_id"] = prompt_df.index

# %%
long_df["response_varies"] = long_df.classification != long_df.classification2
long_df["response_varies"] = np.where(
    long_df.classification2 != long_df.classification3,
    True,
    long_df.response_varies,
)
long_df.response_varies.value_counts()


# %%
long_df.fingerprint.value_counts()
# %%
# Load empty results file (just has prompts sheets)
wb = load_workbook(destination_path)
wb.create_sheet("combo_results")
ws = wb["combo_results"]

ws.cell(row=1, column=1, value="Combo")
ws.cell(row=1, column=2, value="Part 1 ID")
ws.cell(row=1, column=3, value="Part 2 ID")
ws.cell(row=1, column=4, value="Accuracy")
ws.cell(row=1, column=5, value="Precision")
ws.cell(row=1, column=6, value="Recall")
ws.cell(row=1, column=7, value="F1")
ws.cell(row=1, column=8, value="Prompt")

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    part1_id = prompt_df.loc[prompt_df.combo_id == combo, "part1"].values[0]
    part2_id = prompt_df.loc[prompt_df.combo_id == combo, "part2"].values[0]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    ws.cell(row=row, column=1, value=combo)
    ws.cell(row=row, column=2, value=part1_id)
    ws.cell(row=row, column=3, value=part2_id)
    ws.cell(row=row, column=4, value=round(accuracy, 2))
    ws.cell(row=row, column=5, value=round(precision, 2))
    ws.cell(row=row, column=6, value=round(recall, 2))
    ws.cell(row=row, column=7, value=round(f1, 2))
    ws.cell(row=row, column=8, value=combo_df.prompt.iloc[0])
    row = row + 1

    wb.save(destination_path)

# %%

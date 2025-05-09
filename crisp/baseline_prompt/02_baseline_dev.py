###
# Tests three definitions and three tasks.
# Print prompts and performance to destination_path below
###
# %%
from crisp.library import start

PROMPT_FILE = "ncb_variants"
# import prompts from here
source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE + ".xlsx"
# export results here, prompts are copied to this file
destination_path = start.MAIN_DIR + "results/" + PROMPT_FILE + "_results_dev.xlsx"

# previous results
previous_results_path = start.MAIN_DIR + "results/" + PROMPT_FILE + "_results.xlsx"
# prompts and prep results file

# %%
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from tqdm import tqdm
from datetime import datetime


from crisp.library import start
from crisp.library import secrets
from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify
from crisp.library import metric_standard_errors

OPENAI_API_KEY = secrets.OPENAI_API_KEY

client = OpenAI(api_key=OPENAI_API_KEY)
export_results.copy_file_and_paste(
    source_path, destination_path
)  # copy prompts to results folder

# text data
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "dev"]


# %% Import prompts
prompt_df = pd.read_excel(previous_results_path, sheet_name="combo_results")

# %%
max_f1_index = prompt_df["F1"].idxmax()
min_f1_index = prompt_df["F1"].idxmin()

prompt_df = prompt_df[prompt_df.Combo.isin([max_f1_index, min_f1_index])]
# %%
# Create combos
prompt_df["text"] = prompt_df["Prompt"]
prompt_df["combo_id"] = prompt_df["Combo"]

full_prompt_texts = prompt_df["text"].tolist()
# %%
formatted_prompts = []
for prompt_text in full_prompt_texts:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])


# %%
# Test response
prompt = formatted_prompts[1]
text = df.text.iloc[0]
messages = prompt + [{"role": "user", "content": text}]
response = client.chat.completions.create(
    model=start.MODEL, messages=messages, temperature=0, n=1
)
for choice in response.choices:
    print(choice.message.content)
# %%
# %% Get responses (one row per response, with timestamp)
response_rows = []
for loop_num in [1, 2]:
    for prompt in formatted_prompts:
        prompt_text = prompt[0]["content"]
        combo_id = prompt_df[prompt_df.text == prompt_text].combo_id.iloc[0]

        for text, participant_id, study, question, human_code in tqdm(
            zip(df.text, df.participant_id, df.study, df.question, df.human_code),
            total=len(df),
        ):
            timestamp = datetime.now().isoformat()
            messages = prompt + [{"role": "user", "content": text}]
            response = client.chat.completions.create(
                model=start.MODEL,
                messages=messages,
                temperature=0.00001,
                n=5,
                seed=start.SEED,
            )

            for i, choice in enumerate(response.choices):
                cleaned_response = choice.message.content
                classification = classify.create_binary_classification_from_response(
                    cleaned_response
                )

                response_rows.append(
                    {
                        "participant_id": participant_id,
                        "study": study,
                        "question": question,
                        "text": text,
                        "human_code": human_code,
                        "response": cleaned_response,
                        "classification": classification,
                        "prompt": prompt_text,
                        "model": start.MODEL,
                        "fingerprint": response.system_fingerprint,
                        "combo_id": combo_id,
                        "response_number": i + 1,
                        "loop_num": loop_num,
                        "timestamp": timestamp,
                    }
                )

long_df = pd.DataFrame(response_rows)
long_df.to_excel(start.DATA_DIR + "temp/ncb_variants_responses_dev.xlsx", index=False)
# %%
# Load empty results file (just has prompts sheets)
wb = load_workbook(destination_path)
wb.create_sheet("combo_results")
ws = wb["combo_results"]

ws.cell(row=1, column=1, value="Combo")
ws.cell(row=1, column=2, value="Accuracy")
ws.cell(row=1, column=3, value="Precision")
ws.cell(row=1, column=4, value="Recall")
ws.cell(row=1, column=5, value="F1")
ws.cell(row=1, column=6, value="Accuracy SE")
ws.cell(row=1, column=7, value="Precision SE")
ws.cell(row=1, column=8, value="Recall SE")
ws.cell(row=1, column=9, value="F1 SE")
ws.cell(row=1, column=10, value="Prompt")

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    combo_df = long_df[long_df.combo_id == combo]
    combo_df = combo_df[combo_df.response_number == 1]
    combo_df = combo_df[combo_df.loop_num == 1]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    _, accuracy_se = metric_standard_errors.bootstrap_accuracy(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, precision_se = metric_standard_errors.bootstrap_precision(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, recall_se = metric_standard_errors.bootstrap_recall(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, f1_se = metric_standard_errors.bootstrap_f1(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    ws.cell(row=row, column=1, value=combo)
    ws.cell(row=row, column=2, value=round(accuracy, 2))
    ws.cell(row=row, column=3, value=round(precision, 2))
    ws.cell(row=row, column=4, value=round(recall, 2))
    ws.cell(row=row, column=5, value=round(f1, 2))
    ws.cell(row=row, column=6, value=round(accuracy_se, 2))
    ws.cell(row=row, column=7, value=round(precision_se, 2))
    ws.cell(row=row, column=8, value=round(recall_se, 2))
    ws.cell(row=row, column=9, value=round(f1_se, 2))
    ws.cell(row=row, column=10, value=combo_df.prompt.iloc[0])
    row = row + 1

    wb.save(destination_path)

# %%

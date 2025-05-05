# %%
import shutil
import os

import numpy as np
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from tqdm import tqdm


from crisp.library import start
from crisp.library import secrets
from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify
from crisp.library import metric_standard_errors

OPENAI_API_KEY = secrets.OPENAI_API_KEY

MODEL = "gpt-4.1-mini"
client = OpenAI(api_key=OPENAI_API_KEY)

#% 
# Questions:
# Using part2 variant part_num 0. Which part1 variant is the best?
# Then, conditional on that, which part2 variant is the best?
# Then, which part 1 + part2 combo is the best?

df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "train"]

# %% Create list of messages
PROMPT_FILE = "ncb_variants"
source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE + ".xlsx"
destination_path = start.MAIN_DIR + "results/" + PROMPT_FILE + ".xlsx"


part1_variants = pd.read_excel(source_path, sheet_name="part1")
part1_variants = part1_variants.set_index("part_num")
part2_variants = pd.read_excel(source_path, sheet_name="part2")
part2_variants = part2_variants.set_index("part_num")

# %%

part2_partnum0 = part2_variants.loc[0]["prompt_part"] 

full_prompt_texts = []
for part1 in part1_variants["prompt_part"]:
    full_prompt_texts.append(part1 + " " + part2_partnum0 + " ")

formatted_prompts = []
for prompt_text in full_prompt_texts:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])


# Test response
prompt = formatted_prompts[1]
text = df.text.iloc[0]
messages = prompt + [{"role": "user", "content": text}]
response = client.chat.completions.create(
    model=MODEL,
    messages=messages,
    temperature=0,
    n = 1
)
for choice in response.choices:
    print(choice.message.content)
# %%


# %% Get responses
response_dfs = []
combo_id = 0
for prompt in formatted_prompts:
    responses = []
    classifications = []
    for text in tqdm(df.text):
        messages = prompt + [{"role": "user", "content": text}]

        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=0,
    )
        cleaned_response = response.choices[0].message.content

        responses.append(cleaned_response)
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )
        classifications.append(classification)
    response_df = pd.DataFrame(
        {
            "participant_id": df.participant_id,
            "study": df.study,
            "question": df.question,
            "text": df.text,
            "human_code": df.human_code,
            "response": responses,
            "classification": classifications,
            "prompt": prompt[0]["content"],
            "model": MODEL,
            "combo_id": combo_id,
        }
    )
    response_dfs.append(response_df)
    combo_id += 1

long_df = pd.concat(response_dfs)
long_df = long_df.dropna(subset=["classification"])
long_df.to_excel(start.DATA_DIR + "temp/ncb_variants_responses.xlsx", index=False)
# %%
export_results.copy_file_and_paste(source_path, destination_path)
wb = load_workbook(destination_path)
# create new sheet
if "combos" in wb.sheetnames:
    wb.remove(wb["combos"])
wb.create_sheet("combo_results")
ws = wb["combo_results"]
previous_column_count = 1

ws.cell(row=1, column=previous_column_count, value="Combo")
ws.cell(row=1, column=previous_column_count + 1 , value="Accuracy")
ws.cell(row=1, column=previous_column_count + 2, value="Precision")
ws.cell(row=1, column=previous_column_count + 3, value="Recall")
ws.cell(row=1, column=previous_column_count + 4, value="F1")

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    ws.cell(row=row, column=1, value=combo)
    ws.cell(row=row, column=2, value=round(accuracy, 2))
    ws.cell(row=row, column=3, value=round(precision, 2))
    ws.cell(row=row, column=4, value=round(recall, 2))
    ws.cell(row=row, column=5, value=round(f1, 2))
    row = row + 1

    wb.save(destination_path)



# %%
# combo1 = long_df[long_df.combo_id == 1]
# f1_se = metric_standard_errors.bootstrap_f1(
#     combo1.human_code, combo1.classification, n_bootstraps=1000, random_state=12
# )

# # ci around se
# f1_ci = (
#     f1_se[0] - .84 * f1_se[1],
#     f1_se[0] + .84 * f1_se[1],
# )
# %%
# import the results
results_df = pd.read_excel(destination_path, sheet_name="combo_results")
# max f1 - min f1
max_df = results_df.F1.max()
min_df = results_df.F1.min()
diff = max_df - min_df
print(f"Difference between max and min F1: {diff}")

# combo id of the best f1
best_f1_combo = results_df.loc[results_df.F1.idxmax()]["Combo"]
# %%
# Question 2
###
#
###
################
part1_text = part1_variants.loc[best_f1_combo]["prompt_part"]

full_prompt_texts = []
for part2 in part2_variants["prompt_part"]:
    full_prompt_texts.append(part1_text + " " + part2 + " ")

formatted_prompts = []
for prompt_text in full_prompt_texts:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])


# Test response
prompt = formatted_prompts[1]
text = df.text.iloc[0]
messages = prompt + [{"role": "user", "content": text}]
response = client.chat.completions.create(
    model=MODEL,
    messages=messages,
    temperature=0,
    n = 1
)
for choice in response.choices:
    print(choice.message.content)
# %%


# %% Get responses
response_dfs = []
combo_id = 0
for prompt in formatted_prompts:
    responses = []
    classifications = []
    for text in tqdm(df.text):
        messages = prompt + [{"role": "user", "content": text}]

        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=0,
    )
        cleaned_response = response.choices[0].message.content

        responses.append(cleaned_response)
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )
        classifications.append(classification)
    response_df = pd.DataFrame(
        {
            "participant_id": df.participant_id,
            "study": df.study,
            "question": df.question,
            "text": df.text,
            "human_code": df.human_code,
            "response": responses,
            "classification": classifications,
            "prompt": prompt[0]["content"],
            "model": MODEL,
            "combo_id": combo_id,
        }
    )
    response_dfs.append(response_df)
    combo_id += 1

long_df = pd.concat(response_dfs)
long_df = long_df.dropna(subset=["classification"])
long_df.to_excel(start.DATA_DIR + "temp/ncb_variants_part2_responses.xlsx", index=False)
# %%
export_results.copy_file_and_paste(source_path, destination_path)
wb = load_workbook(destination_path)
# create new sheet
if "combos" in wb.sheetnames:
    wb.remove(wb["combos"])
wb.create_sheet("combo_results_part2")
ws = wb["combo_results_part2"]
previous_column_count = 1

ws.cell(row=1, column=previous_column_count, value="Combo")
ws.cell(row=1, column=previous_column_count + 1 , value="Accuracy")
ws.cell(row=1, column=previous_column_count + 2, value="Precision")
ws.cell(row=1, column=previous_column_count + 3, value="Recall")
ws.cell(row=1, column=previous_column_count + 4, value="F1")

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    ws.cell(row=row, column=1, value=combo)
    ws.cell(row=row, column=2, value=round(accuracy, 2))
    ws.cell(row=row, column=3, value=round(precision, 2))
    ws.cell(row=row, column=4, value=round(recall, 2))
    ws.cell(row=row, column=5, value=round(f1, 2))
    row = row + 1

    wb.save(destination_path)



# %%
results_df = pd.read_excel(destination_path, sheet_name="combo_results_part2")
# max f1 - min f1
max_df = results_df.F1.max()
min_df = results_df.F1.min()
diff = max_df - min_df
print(f"Difference between max and min F1: {diff}")

# combo id of the best f1
best_f1_combo = results_df.loc[results_df.F1.idxmax()]["Combo"]
# combo1 = long_df[long_df.combo_id == 1]
# f1_se = metric_standard_errors.bootstrap_f1(
#     combo1.human_code, combo1.classification, n_bootstraps=1000, random_state=12
# )

# # ci around se
# f1_ci = (
#     f1_se[0] - .84 * f1_se[1],
#     f1_se[0] + .84 * f1_se[1],

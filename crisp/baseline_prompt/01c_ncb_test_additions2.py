###
# Test whether additional additions to the prompt improve performance
###


# %%
import shutil
import os

import numpy as np
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from tqdm import tqdm


from crisp.library import start
from crisp.library import secrets
from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify
from crisp.library import metric_standard_errors

OPENAI_API_KEY = secrets.OPENAI_API_KEY

MODEL = "gpt-4.1"
client = OpenAI(api_key=OPENAI_API_KEY)

PART1_ID = 0
PART2_ID = 0
OPT_ID_KEEPER = 2

# %%
# text data
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "train"]
# prompts and prep results file
PROMPT_FILE = "ncb_variants"
source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE + ".xlsx"
destination_path = (
    start.MAIN_DIR + "results/" + PROMPT_FILE + "_additions2_results.xlsx"
)
export_results.copy_file_and_paste(
    source_path, destination_path
)  # copy prompts to results folder

# %% Import prompts
part1_variants = pd.read_excel(source_path, sheet_name="part1")
part1_variants = part1_variants.set_index("part_num")
part2_variants = pd.read_excel(source_path, sheet_name="part2")
part2_variants = part2_variants.set_index("part_num")


# %%
optionals = pd.read_excel(source_path, sheet_name="opt")
optionals = optionals.set_index("part_num")

keeper_text = optionals.loc[OPT_ID_KEEPER]["prompt_part"]
optionals = optionals.drop(OPT_ID_KEEPER)
# %%
# Add one
full_prompt_texts = []
opt_ids = []
for opt, opt_id in zip(optionals["prompt_part"], optionals.index):

    prompt_text = (
        part1_variants.loc[PART1_ID]["prompt_part"]
        + " "
        + keeper_text
        + " "
        + opt
        + " "
        + part2_variants.loc[PART2_ID]["prompt_part"]
        + " "
    )
    full_prompt_texts.append(prompt_text)
    opt_ids.append(opt_id)

# %%
# Create combos


prompt_df = pd.DataFrame(full_prompt_texts, columns=["text"])
prompt_df["opt_id"] = opt_ids

# %%

formatted_prompts = []
for prompt_text in full_prompt_texts:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])


# %%
# Test response
prompt = formatted_prompts[1]
text = df.text.iloc[0]
messages = prompt + [{"role": "user", "content": text}]
response = client.chat.completions.create(
    model=MODEL, messages=messages, temperature=0, n=1
)
for choice in response.choices:
    print(choice.message.content)
# %%
# %% Get responses
response_dfs = []
for prompt, opt_id in zip(formatted_prompts, opt_ids):
    responses = []
    classifications = []
    for text in tqdm(df.text):
        messages = prompt + [{"role": "user", "content": text}]

        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=0,
        )
        cleaned_response = response.choices[0].message.content

        responses.append(cleaned_response)
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )
        classifications.append(classification)
    response_df = pd.DataFrame(
        {
            "participant_id": df.participant_id,
            "study": df.study,
            "question": df.question,
            "text": df.text,
            "human_code": df.human_code,
            "response": responses,
            "classification": classifications,
            "prompt": prompt[0]["content"],
            "model": MODEL,
            "opt_id": opt_id,
        }
    )
    response_dfs.append(response_df)
    opt_id += 1

long_df = pd.concat(response_dfs)
long_df = long_df.dropna(subset=["classification"])
long_df.to_excel(
    start.DATA_DIR + "temp/ncb_variants_plus_opt_responses.xlsx", index=False
)
# %%
# Load empty results file (just has prompts sheets)
wb = load_workbook(destination_path)
wb.create_sheet("combo_results")
ws = wb["combo_results"]

ws.cell(row=1, column=1, value="Opt ID")
ws.cell(row=1, column=2, value="Accuracy")
ws.cell(row=1, column=3, value="Precision")
ws.cell(row=1, column=4, value="Recall")
ws.cell(row=1, column=5, value="F1")

row = 2
for combo in long_df.opt_id.unique():
    combo_df = long_df[long_df.opt_id == combo]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]
    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    ws.cell(row=row, column=1, value=combo)
    ws.cell(row=row, column=2, value=round(accuracy, 2))
    ws.cell(row=row, column=3, value=round(precision, 2))
    ws.cell(row=row, column=4, value=round(recall, 2))
    ws.cell(row=row, column=5, value=round(f1, 2))
    row = row + 1

    wb.save(destination_path)

# %%
KEEPER = 3

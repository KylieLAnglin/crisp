# %%
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
from datetime import datetime
import os

from crisp.library import start, secrets
from crisp.library import format_prompts, classify, metric_standard_errors

CONCEPT = start.CONCEPT
PLATFORM = start.PLATFORM
MODEL = start.MODEL
print(f"Running {CONCEPT} on {PLATFORM} with {MODEL} in train set")
SAMPLE = start.SAMPLE

# ------------------ CONSTANTS ------------------
IMPORT_PROMPT_PATH = start.DATA_DIR + f"prompts/{CONCEPT}_baseline_variants.xlsx"
DATA_PATH = start.DATA_DIR + f"clean/{CONCEPT}.xlsx"

IMPORT_RESULTS_PATH = (
    start.MAIN_DIR + f"results/{PLATFORM}_{CONCEPT}_baseline_zero_results_dev.xlsx"
)

EXPORT_RESPONSE_PATH = (
    start.DATA_DIR
    + f"responses_dev/{PLATFORM}_{CONCEPT}_persona_zero_responses_train.xlsx"
)
EXPORT_RESULTS_PATH = (
    start.MAIN_DIR + f"results/{PLATFORM}_{CONCEPT}_persona_zero_results_train.xlsx"
)
PERSONAS = [
    "You are a brilliant psychologist. ",
    "You are an excellent therapist. ",
    "You are a very smart mental health researcher. ",
]


# ------------------ LOAD DATA ------------------
df = pd.read_excel(DATA_PATH)
df = df[df.split_group == "train"]
if start.SAMPLE:
    df = df.sample(5, random_state=start.SEED)

# ------------------ LOAD PROMPTS ------------------
training_results = pd.read_excel(IMPORT_RESULTS_PATH, sheet_name="results")
tbp_id = training_results.loc[training_results["F1"].idxmax(), "Baseline Prompt ID"]
bbp_id = training_results.loc[training_results["F1"].idxmin(), "Baseline Prompt ID"]

prompt_df = pd.read_excel(IMPORT_PROMPT_PATH, sheet_name="baseline")
prompt_df = prompt_df[prompt_df.baseline_prompt_id.isin([tbp_id, bbp_id])]
prompt_df["prompt"] = prompt_df["prompt"].str.replace("Text: ", "")

top_prompt = prompt_df.loc[prompt_df.baseline_prompt_id == tbp_id, "prompt"].values[0]
bottom_prompt = prompt_df.loc[prompt_df.baseline_prompt_id == bbp_id, "prompt"].values[
    0
]
top_prompt = top_prompt.replace("Text: ", "")
bottom_prompt = bottom_prompt.replace("Text: ", "")
# ------------------ CREATE COMBOS ------------------
combos = []
for persona in PERSONAS:
    for category, base_prompt in [("top", top_prompt), ("bottom", bottom_prompt)]:
        prompt_text = persona + base_prompt
        combo_id = f"{category}_{persona.split()[2].lower()}"
        combos.append(
            {
                "combo_id": combo_id,
                "prompt": prompt_text,
                "category": category,
                "persona": persona,
            }
        )

combo_df = pd.DataFrame(combos)

# ------------------ FORMAT PROMPTS ------------------
# TODO: does this need to be edited for llama? If so,
# add an argument for platform
formatted_prompts = []
for prompt_text in prompt_df["prompt"]:
    message = format_prompts.format_system_message(prompt_text)
    formatted_prompts.append([message])
# ------------------ GENERATE RESPONSES ------------------
response_rows = []
for combo_id, prompt in formatted_prompts.items():
    prompt_text = prompt[0]["content"]  # TODO: is this the same format for llama?
    for text, participant_id, study, question, human_code in tqdm(
        zip(df.text, df.participant_id, df.study, df.question, df.human_code),
        total=len(df),
        desc=f"Prompt: {combo_id}",
    ):
        timestamp = datetime.now().isoformat()
        cleaned_response, system_fingerprint = classify.format_message_and_get_response(
            model_provider=start.PLATFORM,
            prompt=prompt,
            text_to_classify=text,
            temperature=0.0001,
        )
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )

        response_rows.append(
            {
                "participant_id": participant_id,
                "study": study,
                "question": question,
                "text": text,
                "human_code": human_code,
                "response": cleaned_response,
                "classification": classification,
                "prompt": prompt[0]["content"],
                "model": MODEL,
                "prompt_id": combo_id,
                "category": combo_df.loc[
                    combo_df.combo_id == combo_id, "category"
                ].iloc[0],
                "timestamp": timestamp,
                "fingerprint": system_fingerprint,
            }
        )

# ------------------ SAVE RESPONSES ------------------
long_df = pd.DataFrame(response_rows)
long_df.to_excel(EXPORT_RESPONSE_PATH, index=False)

# ------------------ EVALUATE PROMPTS ------------------
if not os.path.exists(EXPORT_RESULTS_PATH):
    wb = Workbook()
    wb.save(EXPORT_RESULTS_PATH)

wb = load_workbook(EXPORT_RESULTS_PATH)
if "results" in wb.sheetnames:
    del wb["results"]
ws = wb.create_sheet("results")

headers = [
    "Combo ID",
    "Baseline Category",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Accuracy SE",
    "Precision SE",
    "Recall SE",
    "F1 SE",
    "Prompt",
]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Filter: loop 1, response 1 only

row = 2
for combo in long_df.combo_id.unique():
    sub_df = long_df[(long_df.combo_id == combo)]
    valid = sub_df.dropna(subset=["human_code", "classification"])
    y_true = valid["human_code"]
    y_pred = valid["classification"]

    accuracy, precision, recall, f1 = classify.print_and_save_metrics(y_true, y_pred)
    _, acc_se = metric_standard_errors.bootstrap_accuracy(y_true, y_pred, 1000, 12)
    _, prec_se = metric_standard_errors.bootstrap_precision(y_true, y_pred, 1000, 12)
    _, rec_se = metric_standard_errors.bootstrap_recall(y_true, y_pred, 1000, 12)
    _, f1_se = metric_standard_errors.bootstrap_f1(y_true, y_pred, 1000, 12)

    results = [
        combo,
        sub_df["category"].iloc[0],
        accuracy,
        precision,
        recall,
        f1,
        acc_se,
        prec_se,
        rec_se,
        f1_se,
        sub_df["prompt"].iloc[0],
    ]

    for col, val in enumerate(results, 1):
        ws.cell(
            row=row, column=col, value=round(val, 2) if isinstance(val, float) else val
        )
    row += 1

wb.save(EXPORT_RESULTS_PATH)
# %%

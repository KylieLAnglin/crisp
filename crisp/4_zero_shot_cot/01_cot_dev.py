# cot_zero_dev.py
# %%
import pandas as pd
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from datetime import datetime
import os
import re

from crisp.library import start, secrets
from crisp.library import format_prompts, classify, metric_standard_errors

SAMPLE = False
# ------------------ CONSTANTS ------------------
OPENAI_API_KEY = secrets.OPENAI_API_KEY
MODEL = start.MODEL
SEED = start.SEED
TEMPERATURE = 0.00001

IMPORT_RESULTS_PATH = start.MAIN_DIR + "results/ncb_variants_results_dev.xlsx"
RESPONSE_PATH = start.DATA_DIR + "responses_dev/ncb_cot_zero_responses_dev.xlsx"
RESULTS_PATH = start.MAIN_DIR + "results/ncb_cot_zero_results_dev.xlsx"

COT_SUFFIX = (
    " First, explain your reasoning step by step. "
    "Then, state your final answer — either Yes or No — using the format: Final Answer: [Yes or No]"
)

client = OpenAI(api_key=OPENAI_API_KEY)


# ------------------ FUNCTION: PARSE FINAL ANSWER ------------------
def parse_prediction_cot(response_text: str) -> int:
    """
    Extract final binary classification from a CoT response using:
    "Final Answer: Yes" or "Final Answer: No"
    """
    match = re.search(
        r"final answer\s*:\s*(yes|no)", response_text.strip(), re.IGNORECASE
    )
    if match:
        return 1 if match.group(1).lower() == "yes" else 0
    return 0  # fallback default


# ------------------ LOAD BASELINE PROMPTS ------------------
baseline_results = pd.read_excel(IMPORT_RESULTS_PATH, sheet_name="results")
top_prompt = baseline_results.loc[baseline_results["F1"].idxmax(), "Prompt"]
bottom_prompt = baseline_results.loc[baseline_results["F1"].idxmin(), "Prompt"]
top_prompt = top_prompt.replace("Text:", "")
bottom_prompt = bottom_prompt.replace("Text:", "")

top_cot = top_prompt + COT_SUFFIX
bottom_cot = bottom_prompt + COT_SUFFIX

prompt_df = pd.DataFrame(
    [
        {"combo_id": "top_cot", "prompt": top_cot},
        {"combo_id": "bottom_cot", "prompt": bottom_cot},
    ]
)

# ------------------ FORMAT PROMPTS ------------------
formatted_prompts = []
prompt_mapping = {}
for row in prompt_df.itertuples():
    msg = format_prompts.format_system_message(row.prompt)
    formatted_prompts.append([msg])
    prompt_mapping[msg["content"]] = row.combo_id

# ------------------ LOAD DEV TEXT DATA ------------------
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "dev"]
df = df[df.text.notna() & df.human_code.notna()]
if SAMPLE:
    df = df.sample(5)
# ------------------ GENERATE RESPONSES ------------------
response_rows = []
for prompt in formatted_prompts:
    prompt_text = prompt[0]["content"]
    combo_id = prompt_mapping[prompt_text]

    for text, participant_id, study, question, human_code in tqdm(
        zip(df.text, df.participant_id, df.study, df.question, df.human_code),
        total=len(df),
        desc=f"{combo_id})",
    ):
        timestamp = datetime.now().isoformat()
        messages = prompt + [{"role": "user", "content": text}]
        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=TEMPERATURE,
            n=1,
            seed=SEED,
        )

        for i, choice in enumerate(response.choices):
            cleaned_response = choice.message.content
            classification = parse_prediction_cot(cleaned_response)

            response_rows.append(
                {
                    "participant_id": participant_id,
                    "study": study,
                    "question": question,
                    "text": text,
                    "human_code": human_code,
                    "response": cleaned_response,
                    "classification": classification,
                    "prompt": prompt_text,
                    "model": MODEL,
                    "fingerprint": response.system_fingerprint,
                    "combo_id": combo_id,
                    "timestamp": timestamp,
                }
            )

# ------------------ SAVE RESPONSES ------------------
long_df = pd.DataFrame(response_rows)
long_df.to_excel(RESPONSE_PATH, index=False)

# ------------------ EVALUATE METRICS ------------------
if not os.path.exists(RESULTS_PATH):
    wb = Workbook()
    wb.save(RESULTS_PATH)

wb = load_workbook(RESULTS_PATH)
if "results" in wb.sheetnames:
    del wb["results"]
ws = wb.create_sheet("results")

headers = [
    "Combo",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Accuracy SE",
    "Precision SE",
    "Recall SE",
    "F1 SE",
    "Prompt",
]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

row = 2

for combo in long_df.combo_id.unique():
    sub_df = long_df[(long_df.combo_id == combo)]
    valid = sub_df.dropna(subset=["human_code", "classification"])
    y_true = valid["human_code"]
    y_pred = valid["classification"]

    accuracy, precision, recall, f1 = classify.print_and_save_metrics(y_true, y_pred)
    _, acc_se = metric_standard_errors.bootstrap_accuracy(y_true, y_pred, 1000)
    _, prec_se = metric_standard_errors.bootstrap_precision(y_true, y_pred, 1000)
    _, rec_se = metric_standard_errors.bootstrap_recall(y_true, y_pred, 1000)
    _, f1_se = metric_standard_errors.bootstrap_f1(y_true, y_pred, 1000)

    prompt_text = sub_df["prompt"].iloc[0]
    results = [
        combo,
        accuracy,
        precision,
        recall,
        f1,
        acc_se,
        prec_se,
        rec_se,
        f1_se,
        prompt_text,
    ]
    for col, val in enumerate(results, 1):
        ws.cell(
            row=row, column=col, value=round(val, 2) if isinstance(val, float) else val
        )
    row += 1

wb.save(RESULTS_PATH)
# %%

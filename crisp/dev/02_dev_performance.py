# %%
import shutil
import os

import numpy as np
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from tqdm import tqdm

from crisp.library import export_results
from crisp.library import format_prompts
from crisp.library import classify
from crisp.library import start
from crisp.library import secrets

# %%

RESULTS_FILE = start.RESULTS_DIR + "performance_ncb_dev.xlsx"

CLASSIFICATION_FILE = start.MAIN_DIR + "data/clean/classifications_ncb_dev.xlsx"
gold = pd.read_excel(CLASSIFICATION_FILE, sheet_name="min")
gold = gold[["unique_text_id", "human_code", "split_group"]]
# %%

# Load the workbook and get the sheet names
wb = load_workbook(CLASSIFICATION_FILE)
sheet_names = wb.sheetnames

wb = load_workbook(RESULTS_FILE)

# Iterate over each sheet
row = 2
start_result_col = 2

for PROMPT_NAME in sheet_names:
    # Skip any non-data sheets
    if PROMPT_NAME.startswith("Sheet"):
        continue

    # Read the gpt classifications from the sheet
    gpt = pd.read_excel(
        CLASSIFICATION_FILE,
        sheet_name=PROMPT_NAME,
    )
    gpt = gpt[["unique_text_id", "classification"]]
    df = gold.merge(gpt, on="unique_text_id", how="inner")
    df["gold_standard"] = df.human_code
    df["gpt_classification"] = df.classification
    df = df[df.split_group == "dev"]

    df = df.dropna()
    accuracy = accuracy_score(df.gold_standard, df.gpt_classification)
    precision = precision_score(df.gold_standard, df.gpt_classification)
    recall = recall_score(df.gold_standard, df.gpt_classification)
    f1 = f1_score(df.gold_standard, df.gpt_classification)

    accuracy_se = np.sqrt((accuracy * (1 - accuracy)) / len(df))
    precision_se = np.sqrt((precision * (1 - precision)) / len(df))
    recall_se = np.sqrt((recall * (1 - recall)) / len(df))
    f1_se = np.sqrt((f1 * (1 - f1)) / len(df))

    ws = wb["results"]

    ws.cell(row=row, column=1, value=PROMPT_NAME)
    col = start_result_col
    for metric in [accuracy, recall, precision, f1]:
        ws.cell(row=row, column=col, value=round(metric, 2))
        col += 1

    row = row + 1
    col = start_result_col
    for metric in [accuracy_se, recall_se, precision_se, f1_se]:
        ws.cell(row=row, column=col, value=f"({round(metric, 2)})")
        col += 1
    row = row + 1
# Save the modified workbook
wb.save(RESULTS_FILE)

# %%

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

# Load the results sheet
RESULTS_FILE = start.RESULTS_DIR + "performance_ncb_dev.xlsx"
df = pd.read_excel(RESULTS_FILE, sheet_name="results", header=None)

# Reformat the data
records = []
i = 1
while i < len(df) - 1:
    print(i)
    row_metrics = df.iloc[i]
    row_se = df.iloc[i + 1]

    prompt = row_metrics.iloc[0]
    accuracy = row_metrics[1]
    recall = row_metrics[2]
    precision = row_metrics[3]

    accuracy_se = float(str(row_se[1]).strip("()"))
    recall_se = float(str(row_se[2]).strip("()"))
    precision_se = float(str(row_se[3]).strip("()"))

    records.append(
        {"Prompt": prompt, "Metric": "Accuracy", "Score": accuracy, "SE": accuracy_se}
    )
    records.append(
        {"Prompt": prompt, "Metric": "Recall", "Score": recall, "SE": recall_se}
    )
    records.append(
        {
            "Prompt": prompt,
            "Metric": "Precision",
            "Score": precision,
            "SE": precision_se,
        }
    )

    i += 1
# %%
metrics_df = pd.DataFrame(records)
metrics_df = metrics_df[metrics_df.Prompt.isnull() == False]

# Plotting
plt.figure(figsize=(12, 6))
prompts = metrics_df["Prompt"].unique()
metrics = ["Accuracy", "Recall", "Precision"]
x = np.arange(len(prompts))  # label locations
width = 0.25

for i, metric in enumerate(metrics):
    subset = metrics_df[metrics_df["Metric"] == metric]
    plt.bar(
        x + i * width,
        subset["Score"],
        width=width,
        yerr=subset["SE"],
        capsize=5,
        label=metric,
    )

plt.xticks(x + width, prompts, rotation=45, ha="right")
plt.ylabel("Score")
plt.ylim(0, 1.1)
plt.title("Accuracy, Precision, and Recall per Prompt with Confidence Intervals")
plt.legend()
plt.tight_layout()
plt.show()
# %%

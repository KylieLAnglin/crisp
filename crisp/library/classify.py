# classify.py
import os
from datetime import datetime
import os
import json
import random
import numpy as np
import pandas as pd
from tqdm import tqdm
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from openai import OpenAI
from openpyxl import Workbook, load_workbook


from crisp.library import secrets, start
from . import metric_standard_errors

# ------------------ CLIENT ------------------
OPENAI_API_KEY = secrets.OPENAI_API_KEY
client = OpenAI(api_key=OPENAI_API_KEY)


# ------------------ PROMPT FORMATTING ------------------
def format_system_message(text):
    """
    Formats a system message with the prompt text followed by 'Text:'.
    Ensures consistency across classification calls.
    """
    return {
        "role": "system",
        "content": text.strip() + " Text:",
    }


# ------------------ RESPONSE FUNCTIONS ------------------
def format_message_and_get_response(
    model_provider, prompt, text_to_classify, temperature=0.0001
):
    """
    Send prompt + text to the model and return the response and fingerprint.
    """
    if model_provider == "openai":
        messages = prompt + [{"role": "user", "content": text_to_classify}]
        response = client.chat.completions.create(
            model=start.MODEL,
            messages=messages,
            temperature=temperature,
            n=1,
            seed=start.SEED,
        )
        cleaned_response = response.choices[0].message.content
        return cleaned_response, response.system_fingerprint

    elif model_provider == "llama":
        # Placeholder: add llama integration
        return "llama_fake_response", "llama_fingerprint"

    else:
        raise ValueError(f"Unsupported model provider: {model_provider}")


def create_binary_classification_from_response(response):
    """
    Classifies response as 1 for 'yes', 0 for 'no', np.nan otherwise.
    """
    response = response.lower()
    if "yes" in response:
        return 1
    elif "no" in response:
        return 0
    return np.nan


# ------------------ PROMPT EVALUATION ------------------
def evaluate_prompt(
    prompt_text, prompt_id, df, platform, temperature=0.0001, parser_fn=None
):
    """
    Applies a prompt to classify a DataFrame of texts and returns a list of response rows.
    Assumes binary classification and expects a 'text' and 'human_code' column in df.
    Optionally accepts a custom parser function to override the default response parsing logic.
    """
    rows = []
    formatted_prompt = format_system_message(prompt_text)
    prompt_content = formatted_prompt["content"]

    for text, participant_id, study, question, human_code in tqdm(
        zip(df.text, df.participant_id, df.study, df.question, df.human_code),
        total=len(df),
        desc=f"Prompt ID: {prompt_id}",
    ):
        timestamp = datetime.now().isoformat()
        cleaned_response, system_fingerprint = format_message_and_get_response(
            model_provider=platform,
            prompt=[formatted_prompt],
            text_to_classify=text,
            temperature=temperature,
        )

        if parser_fn:
            classification = parser_fn(cleaned_response)
        else:
            classification = create_binary_classification_from_response(
                cleaned_response
            )

        rows.append(
            {
                "participant_id": participant_id,
                "study": study,
                "question": question,
                "text": text,
                "human_code": human_code,
                "response": cleaned_response,
                "classification": classification,
                "prompt": prompt_content,
                "model": start.MODEL,
                "fingerprint": system_fingerprint,
                "prompt_id": prompt_id,
                "timestamp": timestamp,
            }
        )

    return rows


# ------------------ METRIC UTILITIES ------------------
def print_and_save_metrics(human_codes, classifications):
    """
    Print and return accuracy, precision, recall, F1.
    """
    accuracy = accuracy_score(human_codes, classifications)
    precision = precision_score(human_codes, classifications)
    recall = recall_score(human_codes, classifications)
    f1 = f1_score(human_codes, classifications)

    print(f"Accuracy: {accuracy}")
    print(f"Precision: {precision}")
    print(f"Recall: {recall}")
    print(f"F1: {f1}")

    return accuracy, precision, recall, f1


# ------------------ PROMPT VARIATION ------------------
def generate_prompt_variants(
    model_provider, base_prompt, metaprompt1, metaprompt2, num_variants
):
    """
    Use meta-instructions to generate prompt variants using a model provider.
    """
    variants = []
    for _ in range(num_variants):
        meta_instructions = metaprompt1 + base_prompt + metaprompt2

        if model_provider == "openai":
            response = client.chat.completions.create(
                model=start.MODEL,
                messages=[{"role": "system", "content": meta_instructions}],
                temperature=1,
            )
            new_prompt = response.choices[0].message.content.strip()

        elif model_provider == "llama":
            # Placeholder for llama support
            new_prompt = "llama_fake_response"

        variants.append(new_prompt)

    return variants


# ------------------ EXCEL EXPORT ------------------
def export_results_to_excel(
    df,
    output_path,
    group_col="prompt_id",
    prompt_col="prompt",
    y_true_col="human_code",
    y_pred_col="classification",
    sheet_name="results",
    include_se=True,
    n_bootstraps=1000,
    random_state=12,
):
    """
    Export classification results (with or without standard errors) to an Excel sheet.

    Parameters:
    - group_col: string or list of column names to group results by.
    """
    from openpyxl import Workbook, load_workbook

    # Ensure output file exists
    if not os.path.exists(output_path):
        wb = Workbook()
        wb.save(output_path)

    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Ensure group_col is a list
    if isinstance(group_col, str):
        group_col = [group_col]

    # Set up header
    headers = group_col + ["Accuracy", "Precision", "Recall", "F1"]
    if include_se:
        headers += ["Accuracy SE", "Precision SE", "Recall SE", "F1 SE"]
    headers.append(prompt_col)

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Group by specified columns
    grouped = df.dropna(subset=[y_true_col, y_pred_col]).groupby(group_col)

    row = 2
    for group_vals, group_df in grouped:
        y_true = group_df[y_true_col]
        y_pred = group_df[y_pred_col]
        prompt_text = group_df[prompt_col].iloc[0]

        acc, prec, rec, f1 = print_and_save_metrics(y_true, y_pred)
        result_row = list(group_vals) if isinstance(group_vals, tuple) else [group_vals]
        result_row += [acc, prec, rec, f1]

        if include_se:
            _, acc_se = metric_standard_errors.bootstrap_accuracy(
                y_true, y_pred, n_bootstraps, random_state
            )
            _, prec_se = metric_standard_errors.bootstrap_precision(
                y_true, y_pred, n_bootstraps, random_state
            )
            _, rec_se = metric_standard_errors.bootstrap_recall(
                y_true, y_pred, n_bootstraps, random_state
            )
            _, f1_se = metric_standard_errors.bootstrap_f1(
                y_true, y_pred, n_bootstraps, random_state
            )
            result_row += [acc_se, prec_se, rec_se, f1_se]

        result_row.append(prompt_text)

        for col, val in enumerate(result_row, 1):
            ws.cell(
                row=row,
                column=col,
                value=round(val, 3) if isinstance(val, float) else val,
            )
        row += 1

    wb.save(output_path)


def evaluate_fewshot_prompt_combinations(
    samples,
    df_eval,
    prompt_dict,
    platform,
    temperature=0.0001,
    prefix="fewshot",
    label_format_fn=None,
    verbose=True,
):
    """
    Evaluate few-shot samples across multiple prompt categories.

    Parameters
    ----------
    samples : list of dict
        Each dict should have keys:
            - "sample_id": unique identifier
            - "num_examples": number of examples
            - "examples": list of dicts with "text" and "label"

    df_eval : pd.DataFrame
        DataFrame of evaluation texts and ground-truth labels.

    prompt_dict : dict
        Mapping from category name to base prompt text (e.g., {"top": ..., "bottom": ...}).

    platform : str
        Platform used for inference (e.g., "gpt-4").

    temperature : float, default=0.0001
        Model temperature setting.

    prefix : str, default="fewshot"
        String used in naming prompt_id (e.g., "top_fewshot_12_n5").

    label_format_fn : function or None
        Optional function to convert binary label to string (e.g., 1 → "Yes", 0 → "No").
        If None, defaults to "Yes"/"No".

    verbose : bool, default=True
        Whether to display a progress bar.

    Returns
    -------
    response_rows : list of dict
        Results from classification, with metadata columns: category, num_examples, prompt_id.
    """

    if label_format_fn is None:
        label_format_fn = lambda label: "Yes" if label == 1 else "No"

    response_rows = []
    iterator = tqdm(samples, desc="Evaluating Few-shot Samples") if verbose else samples

    for sample in iterator:
        sample_id = sample["sample_id"]
        num_examples = sample["num_examples"]
        example_block = "\n".join(
            [
                f'Text: "{ex["text"]}"\nAnswer: {label_format_fn(ex["label"])}'
                for ex in sample["examples"]
            ]
        )

        for category, base_prompt in prompt_dict.items():
            base_prompt_clean = base_prompt.replace("Text:", "")
            full_prompt = (
                f"{base_prompt_clean}\nHere are some examples:\n{example_block}\n\n"
            )
            prompt_id = f"{category}_{prefix}_{sample_id}_n{num_examples}"

            eval_rows = evaluate_prompt(
                prompt_text=full_prompt,
                prompt_id=prompt_id,
                df=df_eval,
                platform=platform,
                temperature=temperature,
            )

            for row in eval_rows:
                row["category"] = category
                row["num_examples"] = num_examples
            response_rows.extend(eval_rows)

    return response_rows

# %%
from crisp.library import start
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

ncb_df = pd.read_excel(start.DATA_DIR + "clean/ncb.xlsx")
mm_df = pd.read_excel(start.DATA_DIR + "clean/mm.xlsx")
gratitude_df = pd.read_excel(start.DATA_DIR + "clean/gratitude.xlsx")
big_df = pd.concat([gratitude_df, ncb_df, mm_df], ignore_index=True)
clinical_df = pd.concat([ncb_df, mm_df], ignore_index=True)
# %%
# split_group and human_code
sample_size_dict = {}
for df, concept in zip([gratitude_df, ncb_df, mm_df], ["gratitude", "ncb", "mm"]):
    sample_size_dict[concept] = {}
    for split in ["train", "dev", "test"]:
        sample_size_dict[concept][split] = {}
        sample_size_dict[concept][split]["total"] = len(df[df.split_group == split])
        sample_size_dict[concept][split]["yes"] = len(
            df[(df.split_group == split) & (df.human_code == 1)]
        )
        sample_size_dict[concept][split]["no"] = len(
            df[(df.split_group == split) & (df.human_code == 0)]
        )
    sample_size_dict[concept]["all"] = {}
    sample_size_dict[concept]["all"]["total"] = len(df)

# %%
RESULTS_FILE = start.RESULTS_DIR + "sample_description.xlsx"
wb = load_workbook(RESULTS_FILE)
ws = wb.active

# gratitude
df = gratitude_df
row = 3
col = 2
for split in ["train", "dev", "test"]:
    # yes
    n = sample_size_dict["gratitude"][split]["yes"]
    ws.cell(row=row, column=col, value=n)
    col += 1
    # no
    n = sample_size_dict["gratitude"][split]["no"]
    ws.cell(row=row, column=col, value=n)
    col += 1
# total
n = sample_size_dict["gratitude"]["all"]["total"]
ws.cell(row=row, column=col, value=n)

# ncb
df = ncb_df
row = 4
col = 2
for split in ["train", "dev", "test"]:

    # yes
    n = sample_size_dict["ncb"][split]["yes"]
    ws.cell(row=row, column=col, value=n)
    col += 1
    # no
    n = sample_size_dict["ncb"][split]["no"]
    ws.cell(row=row, column=col, value=n)
    col += 1
# total
n = sample_size_dict["ncb"]["all"]["total"]
ws.cell(row=row, column=col, value=n)

# mm
df = mm_df
row = 5
col = 2
for split in ["train", "dev", "test"]:
    # yes
    n = sample_size_dict["mm"][split]["yes"]
    ws.cell(row=row, column=col, value=n)
    col += 1
    # no
    n = sample_size_dict["mm"][split]["no"]
    ws.cell(row=row, column=col, value=n)
    col += 1
# total
n = sample_size_dict["mm"]["all"]["total"]
ws.cell(row=row, column=col, value=n)

wb.save(RESULTS_FILE)

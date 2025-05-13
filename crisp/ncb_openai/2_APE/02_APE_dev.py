# A4_ape_dev.py
# %%
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
from datetime import datetime
import os

from crisp.library import start, secrets
from crisp.library import format_prompts, classify, metric_standard_errors

SAMPLE = False
# ------------------ CONSTANTS ------------------
OPENAI_API_KEY = secrets.OPENAI_API_KEY
MODEL = start.MODEL
GENERATION_TEMPERATURE = 0.00001
NUM_RESPONSES = 1  # Only one response per input
SEED = start.SEED

PROMPT_PATH_TOP = start.RESULTS_DIR + "ncb_ape_top_results.csv"
PROMPT_PATH_BOTTOM = start.RESULTS_DIR + "ncb_ape_bottom_results.csv"
RESPONSE_PATH = start.DATA_DIR + "responses_dev/ncb_ape_responses_dev.xlsx"
RESULTS_PATH = start.MAIN_DIR + "results/ncb_ape_results_dev.xlsx"

client = OpenAI(api_key=OPENAI_API_KEY)

# ------------------ LOAD PROMPTS ------------------
prompt_df1 = pd.read_csv(PROMPT_PATH_TOP)
prompt_df1["prompt_id"] = (
    "top"
    + prompt_df1.generation.astype(int).astype(str)
    + "_"
    + prompt_df1.variant_id.astype(int).astype(str)
)
prompt_df1 = prompt_df1.set_index("prompt_id")
max_f1_index1 = prompt_df1["f1_score"].idxmax()
prompt_df1 = prompt_df1.loc[[max_f1_index1]]

prompt_df2 = pd.read_csv(PROMPT_PATH_BOTTOM)
prompt_df2["prompt_id"] = (
    "bottom"
    + prompt_df2.generation.astype(int).astype(str)
    + "_"
    + prompt_df2.variant_id.astype(int).astype(str)
)
prompt_df2 = prompt_df2.set_index("prompt_id")
max_f1_index2 = prompt_df2["f1_score"].idxmax()
prompt_df2 = prompt_df2.loc[[max_f1_index2]]

prompt_df = pd.concat([prompt_df1, prompt_df2])
prompt_df["prompt"] = prompt_df["prompt"].str.replace("Text:", "")
prompt_df["combo_id"] = prompt_df.index

# ------------------ FORMAT PROMPTS ------------------
formatted_prompts = {
    row.combo_id: [format_prompts.format_system_message(row.prompt)]
    for row in prompt_df.itertuples()
}

# ------------------ LOAD TEXT DATA ------------------
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "dev"]
if SAMPLE:
    df = df.sample(5)
# ------------------ COLLECT RESPONSES ------------------
response_rows = []
for combo_id in prompt_df.combo_id:
    prompt = formatted_prompts[combo_id]
    prompt_text = prompt[0]["content"]

    for text, participant_id, study, question, human_code in tqdm(
        zip(df.text, df.participant_id, df.study, df.question, df.human_code),
        total=len(df),
        desc=f"Combo {combo_id}",
    ):
        timestamp = datetime.now().isoformat()
        messages = prompt + [{"role": "user", "content": text}]
        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=GENERATION_TEMPERATURE,
            n=1,
            seed=SEED,
        )

        cleaned_response = response.choices[0].message.content
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )

        response_rows.append(
            {
                "participant_id": participant_id,
                "study": study,
                "question": question,
                "text": text,
                "human_code": human_code,
                "response": cleaned_response,
                "classification": classification,
                "prompt": prompt_text,
                "model": MODEL,
                "fingerprint": response.system_fingerprint,
                "combo_id": combo_id,
                "timestamp": timestamp,
            }
        )

long_df = pd.DataFrame(response_rows)
long_df.to_excel(RESPONSE_PATH, index=False)

# ------------------ EXPORT METRICS ------------------
if not os.path.exists(RESULTS_PATH):
    wb = Workbook()
    wb.save(RESULTS_PATH)

wb = load_workbook(RESULTS_PATH)
if "results" in wb.sheetnames:
    del wb["results"]
ws = wb.create_sheet("results")

headers = [
    "Combo",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Accuracy SE",
    "Precision SE",
    "Recall SE",
    "F1 SE",
    "Prompt",
]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

row = 2
for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    human_codes = combo_df.loc[valid_indices, "human_code"]
    classifications = combo_df.loc[valid_indices, "classification"]

    print(f"Combo {combo}")
    accuracy, precision, recall, f1 = classify.print_and_save_metrics(
        human_codes, classifications
    )
    _, accuracy_se = metric_standard_errors.bootstrap_accuracy(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, precision_se = metric_standard_errors.bootstrap_precision(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, recall_se = metric_standard_errors.bootstrap_recall(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )
    _, f1_se = metric_standard_errors.bootstrap_f1(
        human_codes, classifications, n_bootstraps=1000, random_state=12
    )

    results = [
        combo,
        accuracy,
        precision,
        recall,
        f1,
        accuracy_se,
        precision_se,
        recall_se,
        f1_se,
        combo_df.prompt.iloc[0],
    ]
    for col, val in enumerate(results, 1):
        ws.cell(
            row=row, column=col, value=round(val, 2) if isinstance(val, float) else val
        )

    row += 1

wb.save(RESULTS_PATH)
# %%

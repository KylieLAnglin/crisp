# 02_dev_persona.py
# %%
import pandas as pd
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from datetime import datetime
import os

from crisp.library import start, secrets
from crisp.library import format_prompts, classify, metric_standard_errors

SAMPLE = False
# ------------------ CONSTANTS ------------------
OPENAI_API_KEY = secrets.OPENAI_API_KEY
MODEL = start.MODEL
SEED = start.SEED
TEMPERATURE = 0.00001
NUM_RESPONSES = 1  # Only one response per example

RESULTS_TRAIN_PATH = start.MAIN_DIR + "results/ncb_persona_results_train.xlsx"
RESPONSE_PATH = start.DATA_DIR + "responses_dev/ncb_persona_responses_dev.xlsx"
RESULTS_PATH = start.MAIN_DIR + "results/ncb_persona_results_dev.xlsx"

client = OpenAI(api_key=OPENAI_API_KEY)

# ------------------ LOAD BEST PERSONAS ------------------
results_df = pd.read_excel(RESULTS_TRAIN_PATH, sheet_name="results")
top_combo = (
    results_df[results_df["Baseline Category"] == "top"]
    .sort_values(by="F1", ascending=False)
    .iloc[0]
)
bottom_combo = (
    results_df[results_df["Baseline Category"] == "bottom"]
    .sort_values(by="F1", ascending=False)
    .iloc[0]
)

selected_prompts = pd.DataFrame([top_combo, bottom_combo])
selected_prompts["combo_id"] = ["top_best_persona", "bottom_best_persona"]
selected_prompts["Prompt"] = selected_prompts["Prompt"].str.replace(
    "Text:", "", regex=False
)

# ------------------ FORMAT PROMPTS ------------------
formatted_prompts = {
    row.combo_id: [format_prompts.format_system_message(row.Prompt)]
    for row in selected_prompts.itertuples()
}

# ------------------ LOAD DEV TEXT DATA ------------------
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "dev"]
df = df[df.text.notna() & df.human_code.notna()]
if SAMPLE:
    df = df.sample(5)
# ------------------ GENERATE RESPONSES ------------------
response_rows = []
for combo_id, prompt in formatted_prompts.items():
    prompt_text = prompt[0]["content"]

    for text, participant_id, study, question, human_code in tqdm(
        zip(df.text, df.participant_id, df.study, df.question, df.human_code),
        total=len(df),
        desc=f"{combo_id}",
    ):
        timestamp = datetime.now().isoformat()
        messages = prompt + [{"role": "user", "content": text}]
        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=TEMPERATURE,
            n=NUM_RESPONSES,
            seed=SEED,
        )

        cleaned_response = response.choices[0].message.content
        classification = classify.create_binary_classification_from_response(
            cleaned_response
        )

        response_rows.append(
            {
                "participant_id": participant_id,
                "study": study,
                "question": question,
                "text": text,
                "human_code": human_code,
                "response": cleaned_response,
                "classification": classification,
                "prompt": prompt_text,
                "model": MODEL,
                "fingerprint": response.system_fingerprint,
                "combo_id": combo_id,
                "timestamp": timestamp,
            }
        )

# ------------------ SAVE RESPONSES ------------------
long_df = pd.DataFrame(response_rows)
long_df.to_excel(RESPONSE_PATH, index=False)

# ------------------ EVALUATE PROMPTS ------------------
if not os.path.exists(RESULTS_PATH):
    wb = Workbook()
    wb.save(RESULTS_PATH)

wb = load_workbook(RESULTS_PATH)
if "results" in wb.sheetnames:
    del wb["results"]
ws = wb.create_sheet("results")

headers = [
    "Combo",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Accuracy SE",
    "Precision SE",
    "Recall SE",
    "F1 SE",
    "Prompt",
]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

row = 2
for combo in long_df.combo_id.unique():
    sub_df = long_df[long_df.combo_id == combo]
    valid = sub_df.dropna(subset=["human_code", "classification"])
    y_true = valid["human_code"]
    y_pred = valid["classification"]

    accuracy, precision, recall, f1 = classify.print_and_save_metrics(y_true, y_pred)
    _, acc_se = metric_standard_errors.bootstrap_accuracy(y_true, y_pred, 1000, 12)
    _, prec_se = metric_standard_errors.bootstrap_precision(y_true, y_pred, 1000, 12)
    _, rec_se = metric_standard_errors.bootstrap_recall(y_true, y_pred, 1000, 12)
    _, f1_se = metric_standard_errors.bootstrap_f1(y_true, y_pred, 1000, 12)

    prompt_text = sub_df["prompt"].iloc[0]
    results = [
        combo,
        accuracy,
        precision,
        recall,
        f1,
        acc_se,
        prec_se,
        rec_se,
        f1_se,
        prompt_text,
    ]
    for col, val in enumerate(results, 1):
        ws.cell(
            row=row, column=col, value=round(val, 2) if isinstance(val, float) else val
        )
    row += 1

wb.save(RESULTS_PATH)
# %%

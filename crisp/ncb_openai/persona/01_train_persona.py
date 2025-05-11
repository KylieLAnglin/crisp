# 01_train_persona.py
# %%
import pandas as pd
from openai import OpenAI
from tqdm import tqdm
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

from crisp.library import start, secrets
from crisp.library import classify, format_prompts, metric_standard_errors

# ------------------ CONSTANTS ------------------
OPENAI_API_KEY = secrets.OPENAI_API_KEY
MODEL = start.MODEL
SEED = start.SEED
NUM_RESPONSES = 5
TEMPERATURE = 0.00001

BASELINE_RESULTS_PATH = start.MAIN_DIR + "results/ncb_variants_results_dev.xlsx"
TEXT_DATA_PATH = start.DATA_DIR + "clean/negative_core_beliefs.xlsx"

RESPONSE_PATH = start.DATA_DIR + "responses_train/ncb_persona_responses_train.xlsx"
RESULTS_PATH = start.MAIN_DIR + "results/ncb_persona_results_train.xlsx"

PERSONAS = [
    "You are a brilliant psychologist. ",
    "You are an excellent therapist. ",
    "You are a very smart mental health researcher. ",
]

client = OpenAI(api_key=OPENAI_API_KEY)

# ------------------ LOAD DATA ------------------
df = pd.read_excel(TEXT_DATA_PATH)
df = df[df.split_group == "train"]
df = df[df.text.notna() & df.human_code.notna()]

baseline_results = pd.read_excel(BASELINE_RESULTS_PATH, sheet_name="results")
top_prompt = baseline_results.loc[baseline_results["F1"].idxmax(), "Prompt"]
bottom_prompt = baseline_results.loc[baseline_results["F1"].idxmin(), "Prompt"]

# ------------------ CREATE COMBOS ------------------
combos = []
for persona in PERSONAS:
    for category, base_prompt in [("top", top_prompt), ("bottom", bottom_prompt)]:
        prompt_text = persona + base_prompt
        combo_id = f"{category}_{persona.split()[2].lower()}"
        combos.append(
            {
                "combo_id": combo_id,
                "prompt": prompt_text,
                "category": category,
                "persona": persona,
            }
        )

combo_df = pd.DataFrame(combos)

# ------------------ FORMAT PROMPTS ------------------
formatted_prompts = {}
for row in combo_df.itertuples():
    system_msg = format_prompts.format_system_message(row.prompt)
    formatted_prompts[row.combo_id] = [system_msg]

# ------------------ GENERATE RESPONSES ------------------
response_rows = []
for loop_num in [1, 2]:
    for combo_id, prompt in formatted_prompts.items():
        for text, participant_id, study, question, human_code in tqdm(
            zip(df.text, df.participant_id, df.study, df.question, df.human_code),
            total=len(df),
            desc=f"Prompt: {combo_id} (Loop {loop_num})",
        ):
            timestamp = datetime.now().isoformat()
            messages = prompt + [{"role": "user", "content": text}]
            response = client.chat.completions.create(
                model=MODEL,
                messages=messages,
                temperature=TEMPERATURE,
                n=NUM_RESPONSES,
                seed=SEED,
            )

            for i, choice in enumerate(response.choices):
                cleaned_response = choice.message.content
                classification = classify.create_binary_classification_from_response(
                    cleaned_response
                )
                response_rows.append(
                    {
                        "participant_id": participant_id,
                        "study": study,
                        "question": question,
                        "text": text,
                        "human_code": human_code,
                        "response": cleaned_response,
                        "classification": classification,
                        "prompt": prompt[0]["content"],
                        "combo_id": combo_id,
                        "persona": combo_df.loc[
                            combo_df.combo_id == combo_id, "persona"
                        ].iloc[0],
                        "category": combo_df.loc[
                            combo_df.combo_id == combo_id, "category"
                        ].iloc[0],
                        "response_number": i + 1,
                        "loop_num": loop_num,
                        "timestamp": timestamp,
                        "model": MODEL,
                        "fingerprint": response.system_fingerprint,
                    }
                )

# ------------------ SAVE RESPONSES ------------------
long_df = pd.DataFrame(response_rows)
long_df.to_excel(RESPONSE_PATH, index=False)

# ------------------ EVALUATE PROMPTS ------------------
if not os.path.exists(RESULTS_PATH):
    wb = Workbook()
    wb.save(RESULTS_PATH)

wb = load_workbook(RESULTS_PATH)
if "results" in wb.sheetnames:
    del wb["results"]
ws = wb.create_sheet("results")

headers = [
    "Combo ID",
    "Persona",
    "Baseline Category",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Accuracy SE",
    "Precision SE",
    "Recall SE",
    "F1 SE",
    "Prompt",
]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Filter: loop 1, response 1 only
FILTER = (long_df.response_number == 1) & (long_df.loop_num == 1)

row = 2
for combo in long_df.combo_id.unique():
    sub_df = long_df[(long_df.combo_id == combo) & FILTER]
    valid = sub_df.dropna(subset=["human_code", "classification"])
    y_true = valid["human_code"]
    y_pred = valid["classification"]

    accuracy, precision, recall, f1 = classify.print_and_save_metrics(y_true, y_pred)
    _, acc_se = metric_standard_errors.bootstrap_accuracy(y_true, y_pred, 1000, 12)
    _, prec_se = metric_standard_errors.bootstrap_precision(y_true, y_pred, 1000, 12)
    _, rec_se = metric_standard_errors.bootstrap_recall(y_true, y_pred, 1000, 12)
    _, f1_se = metric_standard_errors.bootstrap_f1(y_true, y_pred, 1000, 12)

    results = [
        combo,
        sub_df["persona"].iloc[0],
        sub_df["category"].iloc[0],
        accuracy,
        precision,
        recall,
        f1,
        acc_se,
        prec_se,
        rec_se,
        f1_se,
        sub_df["prompt"].iloc[0],
    ]

    for col, val in enumerate(results, 1):
        ws.cell(
            row=row, column=col, value=round(val, 2) if isinstance(val, float) else val
        )
    row += 1

wb.save(RESULTS_PATH)
# %%

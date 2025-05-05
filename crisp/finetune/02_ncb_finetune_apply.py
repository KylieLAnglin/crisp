# %%
import shutil
import os

import numpy as np
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score

import json
import jsonlines
from library import start
from library import secrets

OPENAI_API_KEY = secrets.OPENAI_API_KEY

client = OpenAI(api_key=OPENAI_API_KEY)

# %% Text file
df = pd.read_excel(start.DATA_DIR + "clean/negative_core_beliefs.xlsx")
df = df[df.split_group == "dev"]

# %%
MODEL = "ft:gpt-4o-2024-08-06:personal:ncb-2025-03-18:BCTn6Wab"


# %% Copy prompt file to export results
PROMPT_FILE = "ncb_finetune_prompt.xlsx"

source_path = start.DATA_DIR + "prompts/" + PROMPT_FILE
destination_path = start.MAIN_DIR + "results/" + PROMPT_FILE
os.makedirs(os.path.dirname(destination_path), exist_ok=True)
shutil.copy2(source_path, destination_path)

# %% Import prompts
prompt_components = pd.read_excel(
    start.DATA_DIR + "prompts/" + PROMPT_FILE, sheet_name="components"
)
prompt_components = prompt_components.set_index("component")
prompt_combos = pd.read_excel(
    start.DATA_DIR + "prompts/" + PROMPT_FILE, sheet_name="combos"
)
prompt_combos = prompt_combos.set_index("combo_id")

prompts = []
for i, row in prompt_combos.iterrows():
    prompt = {}
    text = ""
    for component in row:
        if pd.isna(component):
            continue
        text += prompt_components.loc[component, "text"]
    prompt["text"] = text
    prompt["combo_id"] = i
    prompts.append(prompt)

prompts_df = pd.DataFrame(prompts)
# %% Format prompts
formatted_prompt = []

for i, row in prompts_df.iterrows():
    formatted_prompt.append(
        {
            "role": "system",
            "content": row["text"],
        }
    )

# %% Get responses
response_dfs = []
combo_id = 1
for prompt in formatted_prompt:
    responses = []
    classifications = []
    for text in df.text:
        messages = [prompt] + [{"role": "user", "content": text}]

        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            temperature=0.00,
        )
        cleaned_response = response.choices[0].message.content
        responses.append(cleaned_response)
        if "yes" in cleaned_response.lower():
            classification = 1
        elif "no" in cleaned_response.lower():
            classification = 0
        else:
            classification = np.nan
        classifications.append(classification)
    response_df = pd.DataFrame(
        {
            "participant_id": df.participant_id,
            "study": df.study,
            "question": df.question,
            "text": df.text,
            "human_code": df.human_code,
            "response": responses,
            "classification": classifications,
            "combo_id": combo_id,
        }
    )
    response_dfs.append(response_df)
    combo_id += 1

long_df = pd.concat(response_dfs)
long_df = long_df.dropna(subset=["classification"])

# %%
# add accuracy, precision, recall, f1 columns to the combos sheet in results
wb = load_workbook(destination_path)
ws = wb["combos"]

column_count = ws.max_column
col = ws.max_column + 1

ws.cell(row=1, column=col, value="Accuracy")
ws.cell(row=1, column=col + 1, value="Precision")
ws.cell(row=1, column=col + 2, value="Recall")
ws.cell(row=1, column=col + 3, value="F1")

for combo in long_df.combo_id.unique():
    combo_df = long_df[long_df.combo_id == combo]
    valid_indices = combo_df.dropna(subset=["human_code", "classification"]).index
    accuracy = accuracy_score(
        combo_df.loc[valid_indices, "human_code"],
        combo_df.loc[valid_indices, "classification"],
    )
    precision = precision_score(
        combo_df.loc[valid_indices, "human_code"],
        combo_df.loc[valid_indices, "classification"],
        zero_division=0,
    )
    recall = recall_score(
        combo_df.loc[valid_indices, "human_code"],
        combo_df.loc[valid_indices, "classification"],
    )
    f1 = f1_score(
        combo_df.loc[valid_indices, "human_code"],
        combo_df.loc[valid_indices, "classification"],
    )
    ws.cell(row=combo + 1, column=col, value=round(accuracy, 2))
    ws.cell(row=combo + 1, column=col + 1, value=round(precision, 2))
    ws.cell(row=combo + 1, column=col + 2, value=round(recall, 2))
    ws.cell(row=combo + 1, column=col + 3, value=round(f1, 2))
    print(f"Combo {combo}")
    print(f"Accuracy: {accuracy}")
    print(f"Precision: {precision}")
    print(f"Recall: {recall}")
    print(f"F1: {f1}")

wb.save(destination_path)
# %%
